import os
import re
import io
import tempfile
import logging
from datetime import datetime
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash, session, abort
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_login import LoginManager, login_required, login_user, logout_user, current_user, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# пути для стабильной работы на любом хостинге
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(BASE_DIR)
TEMPLATE_DIR = os.path.join(PROJECT_ROOT, 'frontend', 'templates')
STATIC_DIR = os.path.join(PROJECT_ROOT, 'frontend', 'static')

# нормализация путей
TEMPLATE_DIR = os.path.abspath(TEMPLATE_DIR)
STATIC_DIR = os.path.abspath(STATIC_DIR)

# создаём папку для базы данных
INSTANCE_DIR = os.path.join(BASE_DIR, 'instance')
os.makedirs(INSTANCE_DIR, exist_ok=True)

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)

# конфигурация
DB_PATH = os.path.join(INSTANCE_DIR, 'karman_prorab.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'change-this-in-production-secret-key-2024')
app.config['PERMANENT_SESSION_LIFETIME'] = 30 * 24 * 3600  # 30 дней
app.config['SESSION_COOKIE_SECURE'] = False  # Для HTTP (True для HTTPS)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# настройки загрузки файлов
UPLOAD_FOLDER_RECEIPTS = os.path.join(STATIC_DIR, 'uploads', 'receipts')
UPLOAD_FOLDER_PHOTOS = os.path.join(STATIC_DIR, 'uploads', 'project_photos')
UPLOAD_FOLDER_LOGO = os.path.join(STATIC_DIR, 'uploads', 'logo')
UPLOAD_FOLDER_TEMPLATES = os.path.join(STATIC_DIR, 'uploads', 'templates')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'xlsx', 'csv'}

app.config['UPLOAD_FOLDER_RECEIPTS'] = UPLOAD_FOLDER_RECEIPTS
app.config['UPLOAD_FOLDER_PHOTOS'] = UPLOAD_FOLDER_PHOTOS
app.config['UPLOAD_FOLDER_LOGO'] = UPLOAD_FOLDER_LOGO
app.config['UPLOAD_FOLDER_TEMPLATES'] = UPLOAD_FOLDER_TEMPLATES
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# создаём папки для загрузок
os.makedirs(UPLOAD_FOLDER_RECEIPTS, exist_ok=True)
os.makedirs(UPLOAD_FOLDER_PHOTOS, exist_ok=True)
os.makedirs(UPLOAD_FOLDER_LOGO, exist_ok=True)
os.makedirs(UPLOAD_FOLDER_TEMPLATES, exist_ok=True)

# инициализация расширений
db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Пожалуйста, войдите в систему'

# константы
PROJECT_STATUSES = ['На этапе согласования', 'В работе', 'Завершённые', 'Отложенные']
ESTIMATE_MODES = ['client_no_materials', 'client_with_materials', 'internal']

# проверка WeasyPrint
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
    logger.info("WeasyPrint доступен")
except ImportError as e:
    WEASYPRINT_AVAILABLE = False
    logger.warning(f"WeasyPrint не доступен: {e}")
    HTML = None

# вспомогательные функции
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def login_required_with_message(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Пожалуйста, войдите в систему для доступа к этой странице', 'warning')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# модели бд 

class User(db.Model, UserMixin):
    __tablename__ = 'user'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    full_name = db.Column(db.String(200), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    theme = db.Column(db.String(10), default='light')
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class LegalEntity(db.Model):
    __tablename__ = 'legal_entity'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    entity_type = db.Column(db.String(20), nullable=False)
    name = db.Column(db.String(200))
    inn = db.Column(db.String(12), unique=True)
    ogrn = db.Column(db.String(15))
    address = db.Column(db.String(300))
    logo_filename = db.Column(db.String(200), nullable=True)
    is_default = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref='legal_entities')

class Service(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    unit = db.Column(db.String(20), nullable=False)
    work_price = db.Column(db.Float, nullable=False, default=0.0)
    material_price = db.Column(db.Float, nullable=False, default=0.0)

class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    address = db.Column(db.String(200))
    status = db.Column(db.String(20), default='В работе')
    client_name = db.Column(db.String(100))
    client_phone = db.Column(db.String(50))
    actual_end_date = db.Column(db.String(20))
    estimate_mode = db.Column(db.String(30), default='client_no_materials')
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    legal_entity_id = db.Column(db.Integer, db.ForeignKey('legal_entity.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    user = db.relationship('User', backref='projects')
    legal_entity = db.relationship('LegalEntity', backref='projects')

class ProjectTimeline(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False, unique=True)
    start_date = db.Column(db.String(20))
    end_date = db.Column(db.String(20))
    project = db.relationship('Project', backref=db.backref('timeline', uselist=False))

class ProjectWork(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    service_id = db.Column(db.Integer, db.ForeignKey('service.id'), nullable=False)
    quantity = db.Column(db.Float, nullable=False, default=1)
    total_price = db.Column(db.Float, nullable=False, default=0)
    custom_work_price = db.Column(db.Float, nullable=True)
    custom_material_price = db.Column(db.Float, nullable=True)
    custom_total_price = db.Column(db.Float, nullable=True)
    custom_name = db.Column(db.String(200), nullable=True)
    project = db.relationship('Project', backref='works')
    service = db.relationship('Service')

class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False, default=0)
    project = db.relationship('Project', backref='expenses')

class ExpensePhoto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    expense_id = db.Column(db.Integer, db.ForeignKey('expense.id'), nullable=False)
    filename = db.Column(db.String(200), nullable=False)
    filepath = db.Column(db.String(300), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    expense = db.relationship('Expense', backref='photos')

class ProjectPhoto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    filename = db.Column(db.String(200), nullable=False)
    filepath = db.Column(db.String(300), nullable=False)
    photo_type = db.Column(db.String(20), default='progress')
    description = db.Column(db.String(500))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    project = db.relationship('Project', backref='photos')

class Purchase(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    quantity = db.Column(db.Integer, default=1)
    is_purchased = db.Column(db.Boolean, default=False)
    project = db.relationship('Project', backref='purchases')

class ProductCategory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category_id = db.Column(db.Integer, db.ForeignKey('product_category.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    unit = db.Column(db.String(20), nullable=False)
    price = db.Column(db.Float, nullable=False)
    description = db.Column(db.String(500))
    category = db.relationship('ProductCategory', backref=db.backref('products', lazy=True))

class Setting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.String(500), nullable=True)

    @staticmethod
    def get(key, default=''):
        setting = Setting.query.filter_by(key=key).first()
        return setting.value if setting else default

    @staticmethod
    def set(key, value):
        setting = Setting.query.filter_by(key=key).first()
        if setting:
            setting.value = value
        else:
            setting = Setting(key=key, value=value)
            db.session.add(setting)
        db.session.commit()

class Template(db.Model):
    __tablename__ = 'template'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    filename = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref='templates')

# контектсные процессоры

@app.context_processor
def inject_globals():
    return {
        'project_statuses': PROJECT_STATUSES, 
        'estimate_modes': ESTIMATE_MODES,
        'LegalEntity': LegalEntity,
        'weasyprint_available': WEASYPRINT_AVAILABLE
    }

@app.template_filter('format_number')
def format_number(value):
    if value is None:
        return '0'
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        formatted = f"{value:.2f}".rstrip('0').rstrip('.')
        return formatted
    return str(value)

@app.template_filter('ru_date')
def ru_date(date_str):
    if not date_str:
        return ''
    parts = str(date_str).split('-')
    if len(parts) == 3:
        return f"{parts[2]}.{parts[1]}.{parts[0]}"
    return date_str

def normalize_date(date_str):
    if not date_str:
        return None
    if re.match(r'\d{4}-\d{2}-\d{2}', date_str):
        return date_str
    try:
        return datetime.strptime(date_str, '%d.%m.%Y').strftime('%Y-%m-%d')
    except:
        pass
    return date_str

def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d')
    except:
        pass
    try:
        return datetime.strptime(value, '%d.%m.%Y')
    except:
        pass
    return None

def calculate_progress(project):
    timeline = ProjectTimeline.query.filter_by(project_id=project.id).first()
    if not timeline or not timeline.start_date or not timeline.end_date:
        return 0
    start = parse_date(timeline.start_date)
    end = parse_date(timeline.end_date)
    if not start or not end:
        return 0
    today = datetime.now()
    if today >= end:
        return 100
    if today <= start:
        return 0
    total_days = (end - start).days
    days_passed = (today - start).days
    return min(100, int((days_passed / total_days) * 100)) if total_days > 0 else 0

def get_progress_color(progress):
    if progress < 30: return '#dc3545'
    if progress < 50: return '#fd7e14'
    if progress < 70: return '#ffc107'
    if progress < 90: return '#20c997'
    return '#198754'

# HEALTH CHECK

@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()}), 200

# маршруты аутентификации

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            login_user(user, remember=True)
            session.permanent = True
            next_page = request.args.get('next')
            flash(f'Добро пожаловать, {user.full_name}!', 'success')
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('Неверный email или пароль', 'danger')
    
    return render_template('auth/login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        full_name = request.form.get('full_name')
        phone = request.form.get('phone')
        
        if not all([email, password, full_name, phone]):
            flash('Заполните все поля', 'danger')
            return redirect(url_for('register'))
        
        if password != confirm_password:
            flash('Пароли не совпадают', 'danger')
            return redirect(url_for('register'))
        
        if len(password) < 6:
            flash('Пароль должен быть не менее 6 символов', 'danger')
            return redirect(url_for('register'))
        
        if User.query.filter_by(email=email).first():
            flash('Пользователь с таким email уже существует', 'danger')
            return redirect(url_for('register'))
        
        user = User(email=email, full_name=full_name, phone=phone)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        
        login_user(user, remember=True)
        
        flash('Регистрация прошла успешно! Теперь добавьте данные вашей компании', 'success')
        return redirect(url_for('add_legal_entity'))
    
    return render_template('auth/register.html')

@app.route('/add-legal-entity', methods=['GET', 'POST'])
@login_required
def add_legal_entity():
    if request.method == 'POST':
        entity_type = request.form.get('entity_type')
        inn = request.form.get('inn')
        ogrn = request.form.get('ogrn')
        address = request.form.get('address')
        
        if LegalEntity.query.filter_by(inn=inn, user_id=current_user.id).first():
            flash('Юридическое лицо с таким ИНН уже существует', 'danger')
            return redirect(url_for('add_legal_entity'))
        
        if entity_type == 'ooo':
            name = request.form.get('company_name')
        elif entity_type == 'ip':
            name = request.form.get('full_name')
        else:
            name = request.form.get('full_name')
        
        legal_entity = LegalEntity(
            user_id=current_user.id,
            entity_type=entity_type,
            name=name,
            inn=inn,
            ogrn=ogrn or None,
            address=address,
            is_default=LegalEntity.query.filter_by(user_id=current_user.id).count() == 0
        )
        db.session.add(legal_entity)
        db.session.commit()
        
        session['current_entity_id'] = legal_entity.id
        
        flash('Данные компании добавлены!', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('auth/add_legal_entity.html')

# профиль 

@app.route('/profile')
@login_required
def profile():
    return render_template('auth/profile.html')

@app.route('/profile/edit', methods=['GET', 'POST'])
@login_required
def edit_profile():
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if not all([full_name, email, phone]):
            flash('Заполните все поля', 'danger')
            return redirect(url_for('edit_profile'))
        
        if email != current_user.email:
            if User.query.filter_by(email=email).first():
                flash('Email уже используется', 'danger')
                return redirect(url_for('edit_profile'))
        
        current_user.full_name = full_name
        current_user.email = email
        current_user.phone = phone
        
        if new_password:
            if len(new_password) < 6:
                flash('Пароль должен быть не менее 6 символов', 'danger')
                return redirect(url_for('edit_profile'))
            if new_password != confirm_password:
                flash('Пароли не совпадают', 'danger')
                return redirect(url_for('edit_profile'))
            current_user.set_password(new_password)
        
        db.session.commit()
        flash('Профиль успешно обновлён', 'success')
        return redirect(url_for('profile'))
    
    return render_template('auth/edit_profile.html')

@app.route('/profile/delete', methods=['POST'])
@login_required
def delete_profile():
    password = request.form.get('password')
    if not current_user.check_password(password):
        flash('Неверный пароль', 'danger')
        return redirect(url_for('profile'))
    
    # каскадное удаление
    for project in current_user.projects:
        ProjectWork.query.filter_by(project_id=project.id).delete()
        Expense.query.filter_by(project_id=project.id).delete()
        Purchase.query.filter_by(project_id=project.id).delete()
        ProjectPhoto.query.filter_by(project_id=project.id).delete()
        ProjectTimeline.query.filter_by(project_id=project.id).delete()
    
    Project.query.filter_by(user_id=current_user.id).delete()
    LegalEntity.query.filter_by(user_id=current_user.id).delete()
    Template.query.filter_by(user_id=current_user.id).delete()
    
    db.session.delete(current_user)
    db.session.commit()
    
    logout_user()
    flash('Аккаунт удалён', 'info')
    return redirect(url_for('index'))

@app.route('/toggle-theme')
@login_required
def toggle_theme():
    current_user.theme = 'dark' if current_user.theme == 'light' else 'light'
    db.session.commit()
    return redirect(request.referrer or url_for('dashboard'))

# юр. лица 

@app.route('/legal-entities')
@login_required
def legal_entities():
    entities = LegalEntity.query.filter_by(user_id=current_user.id).all()
    return render_template('auth/legal_entities.html', entities=entities)

@app.route('/legal-entity/set-default/<int:entity_id>')
@login_required
def set_default_entity(entity_id):
    LegalEntity.query.filter_by(user_id=current_user.id).update({'is_default': False})
    entity = LegalEntity.query.get_or_404(entity_id)
    if entity.user_id != current_user.id:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('legal_entities'))
    entity.is_default = True
    db.session.commit()
    flash('Юридическое лицо установлено по умолчанию', 'success')
    return redirect(url_for('legal_entities'))

@app.route('/legal-entity/upload-logo/<int:entity_id>', methods=['POST'])
@login_required
def upload_logo(entity_id):
    entity = LegalEntity.query.get_or_404(entity_id)
    if entity.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    if 'logo' not in request.files:
        return jsonify({'success': False, 'message': 'Файл не выбран'})
    
    file = request.files['logo']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Файл не выбран'})
    
    if file and allowed_file(file.filename):
        original_name = secure_filename(file.filename)
        filename = f"logo_{entity_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{original_name}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER_LOGO'], filename)
        file.save(filepath)
        
        if entity.logo_filename:
            old_path = os.path.join(app.config['UPLOAD_FOLDER_LOGO'], entity.logo_filename)
            if os.path.exists(old_path):
                os.remove(old_path)
        
        entity.logo_filename = filename
        db.session.commit()
        
        return jsonify({'success': True, 'logo_url': url_for('static', filename=f'uploads/logo/{filename}')})
    
    return jsonify({'success': False, 'message': 'Недопустимый формат файла'})

@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    flash('Вы вышли из системы', 'info')
    return redirect(url_for('index'))

@app.route('/switch-entity/<int:entity_id>')
@login_required
def switch_entity(entity_id):
    entity = LegalEntity.query.get_or_404(entity_id)
    if entity.user_id != current_user.id:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('dashboard'))
    
    session['current_entity_id'] = entity_id
    flash(f'Вы работаете от имени: {entity.name}', 'success')
    return redirect(request.referrer or url_for('dashboard'))

# основные маршруты 

@app.route('/dashboard')
@login_required
def dashboard():
    projects = Project.query.filter_by(user_id=current_user.id).order_by(Project.created_at.desc()).all()
    
    # статистика
    total_projects = len(projects)
    active_projects = sum(1 for p in projects if p.status == 'В работе')
    completed_projects = sum(1 for p in projects if p.status == 'Завершённые')
    
    # доходы по месяцам
    monthly_data = {}
    for p in projects:
        if p.timeline and p.timeline.start_date:
            try:
                month = p.timeline.start_date[:7]
                total = sum(w.total_price for w in p.works)
                monthly_data[month] = monthly_data.get(month, 0) + total
            except:
                pass
    
    months = sorted(monthly_data.keys())[-6:] if monthly_data else []
    incomes = [monthly_data.get(m, 0) for m in months]
    
    return render_template('dashboard.html', 
                         projects=projects[:5],  # последние 5 проектов
                         total_projects=total_projects,
                         active_projects=active_projects,
                         completed_projects=completed_projects,
                         chart_months=months, 
                         chart_incomes=incomes)

# услуги (работы)

@app.route('/services')
@login_required
def services():
    services = Service.query.order_by(Service.name).all()
    return render_template('services.html', services=services)

@app.route('/services/add', methods=['GET', 'POST'])
@login_required
def add_service():
    if request.method == 'POST':
        try:
            name = request.form.get('name')
            unit = request.form.get('unit')
            work_price = float(request.form.get('work_price', 0))
            material_price = float(request.form.get('material_price', 0))
            
            if not name or not unit:
                flash('Название и единица измерения обязательны', 'danger')
                return redirect(url_for('add_service'))
            
            if work_price < 0 or material_price < 0:
                flash('Цены не могут быть отрицательными', 'danger')
                return redirect(url_for('add_service'))
            
            service = Service(name=name, unit=unit, work_price=work_price, material_price=material_price)
            db.session.add(service)
            db.session.commit()
            flash('Услуга добавлена', 'success')
            return redirect(url_for('services'))
        except Exception as e:
            flash(f'Ошибка: {str(e)}', 'danger')
    
    return render_template('add_service.html')

@app.route('/services/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_service(id):
    service = Service.query.get_or_404(id)
    if request.method == 'POST':
        try:
            service.name = request.form['name']
            service.unit = request.form['unit']
            service.work_price = float(request.form.get('work_price', 0))
            service.material_price = float(request.form.get('material_price', 0))
            
            if service.work_price < 0 or service.material_price < 0:
                flash('Цены не могут быть отрицательными', 'danger')
                return redirect(url_for('edit_service', id=id))
            
            db.session.commit()
            flash('Услуга обновлена', 'success')
            return redirect(url_for('services'))
        except Exception as e:
            flash(f'Ошибка: {str(e)}', 'danger')
    
    return render_template('edit_service.html', service=service)

@app.route('/services/delete/<int:id>')
@login_required
def delete_service(id):
    service = Service.query.get_or_404(id)
    # Проверяем, не используется ли услуга в проектах
    if ProjectWork.query.filter_by(service_id=id).first():
        flash('Нельзя удалить услугу, которая используется в проектах', 'danger')
        return redirect(url_for('services'))
    
    db.session.delete(service)
    db.session.commit()
    flash('Услуга удалена', 'success')
    return redirect(url_for('services'))

# категории и продукты 

@app.route('/categories')
@login_required
def categories():
    categories = ProductCategory.query.order_by(ProductCategory.name).all()
    return render_template('categories.html', categories=categories)

@app.route('/categories/add', methods=['GET', 'POST'])
@login_required
def add_category():
    if request.method == 'POST':
        name = request.form.get('name')
        if not name:
            flash('Название категории обязательно', 'danger')
            return redirect(url_for('add_category'))
        
        category = ProductCategory(name=name)
        db.session.add(category)
        db.session.commit()
        flash('Категория добавлена', 'success')
        return redirect(url_for('categories'))
    
    return render_template('add_category.html')

@app.route('/categories/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_category(id):
    category = ProductCategory.query.get_or_404(id)
    if request.method == 'POST':
        category.name = request.form['name']
        db.session.commit()
        flash('Категория обновлена', 'success')
        return redirect(url_for('categories'))
    
    return render_template('edit_category.html', category=category)

@app.route('/categories/delete/<int:id>')
@login_required
def delete_category(id):
    category = ProductCategory.query.get_or_404(id)
    if Product.query.filter_by(category_id=id).first():
        flash('Нельзя удалить категорию, в которой есть товары', 'danger')
        return redirect(url_for('categories'))
    
    db.session.delete(category)
    db.session.commit()
    flash('Категория удалена', 'success')
    return redirect(url_for('categories'))

@app.route('/products')
@login_required
def products():
    products_list = Product.query.all()
    categories = ProductCategory.query.all()
    return render_template('products.html', products=products_list, categories=categories)

@app.route('/products/by-category/<int:category_id>')
@login_required
def products_by_category(category_id):
    products_list = Product.query.filter_by(category_id=category_id).all()
    return jsonify([{'id': p.id, 'name': p.name, 'unit': p.unit, 'price': p.price} for p in products_list])

@app.route('/products/add', methods=['GET', 'POST'])
@login_required
def add_product():
    if request.method == 'POST':
        try:
            category_id = int(request.form['category_id'])
            name = request.form['name']
            unit = request.form['unit']
            price = float(request.form['price'])
            description = request.form.get('description', '')
            
            if price < 0:
                flash('Цена не может быть отрицательной', 'danger')
                return redirect(url_for('add_product'))
            
            product = Product(
                category_id=category_id,
                name=name,
                unit=unit,
                price=price,
                description=description
            )
            db.session.add(product)
            db.session.commit()
            flash('Продукт добавлен', 'success')
            return redirect(url_for('products'))
        except Exception as e:
            flash(f'Ошибка: {str(e)}', 'danger')
    
    categories = ProductCategory.query.all()
    return render_template('add_product.html', categories=categories)

@app.route('/products/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_product(id):
    product = Product.query.get_or_404(id)
    if request.method == 'POST':
        product.category_id = int(request.form['category_id'])
        product.name = request.form['name']
        product.unit = request.form['unit']
        product.price = float(request.form['price'])
        product.description = request.form.get('description', '')
        
        if product.price < 0:
            flash('Цена не может быть отрицательной', 'danger')
            return redirect(url_for('edit_product', id=id))
        
        db.session.commit()
        flash('Продукт обновлён', 'success')
        return redirect(url_for('products'))
    
    categories = ProductCategory.query.all()
    return render_template('edit_product.html', product=product, categories=categories)

@app.route('/products/delete/<int:id>')
@login_required
def delete_product(id):
    product = Product.query.get_or_404(id)
  
    db.session.delete(product)
    db.session.commit()
    flash('Продукт удалён', 'success')
    return redirect(url_for('products'))

# проекты

@app.route('/projects')
@login_required
def projects():
    projects_list = Project.query.filter_by(user_id=current_user.id).order_by(Project.created_at.desc()).all()
    for p in projects_list:
        p.progress = calculate_progress(p)
    return render_template('projects.html', projects=projects_list)

@app.route('/projects/add', methods=['GET', 'POST'])
@login_required
def add_project():
    if request.method == 'POST':
        try:
            project = Project(
                name=request.form['name'],
                address=request.form.get('address', ''),
                status=request.form.get('status', 'В работе'),
                client_name=request.form.get('client_name', ''),
                client_phone=request.form.get('client_phone', ''),
                estimate_mode=request.form.get('estimate_mode', 'client_no_materials'),
                user_id=current_user.id,
                legal_entity_id=session.get('current_entity_id')
            )
            db.session.add(project)
            db.session.commit()
            
            start_date = normalize_date(request.form.get('start_date'))
            end_date = normalize_date(request.form.get('end_date'))
            if start_date or end_date:
                timeline = ProjectTimeline(
                    project_id=project.id,
                    start_date=start_date,
                    end_date=end_date
                )
                db.session.add(timeline)
                db.session.commit()
            
            flash('Проект создан', 'success')
            return redirect(url_for('project_detail', project_id=project.id))
        except Exception as e:
            flash(f'Ошибка: {str(e)}', 'danger')
    
    return render_template('add_project.html')

@app.route('/projects/delete/<int:id>')
@login_required
def delete_project(id):
    project = Project.query.get_or_404(id)
    if project.user_id != current_user.id:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('projects'))
    
    # каскадное удаление
    ProjectWork.query.filter_by(project_id=id).delete()
    Expense.query.filter_by(project_id=id).delete()
    Purchase.query.filter_by(project_id=id).delete()
    ProjectPhoto.query.filter_by(project_id=id).delete()
    ProjectTimeline.query.filter_by(project_id=id).delete()
    
    db.session.delete(project)
    db.session.commit()
    flash('Проект удалён', 'success')
    return redirect(url_for('projects'))

@app.route('/project/<int:project_id>')
@login_required
def project_detail(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('projects'))
    
    progress = calculate_progress(project)
    works = ProjectWork.query.filter_by(project_id=project_id).all()
    expenses = Expense.query.filter_by(project_id=project_id).all()
    purchases = Purchase.query.filter_by(project_id=project_id).all()
    photos = ProjectPhoto.query.filter_by(project_id=project_id).order_by(ProjectPhoto.uploaded_at.desc()).all()
    timeline = ProjectTimeline.query.filter_by(project_id=project_id).first()
    
    total = sum(w.total_price for w in works)
    total_expenses = sum(e.amount for e in expenses)
    profit = total - total_expenses
    
    services_list = Service.query.order_by(Service.name).all()
    categories = ProductCategory.query.all()
    
    return render_template('project_detail.html',
                         project=project,
                         services=services_list,
                         categories=categories,
                         works=works,
                         expenses=expenses,
                         purchases=purchases,
                         photos=photos,
                         timeline=timeline,
                         progress=progress,
                         progress_color=get_progress_color(progress),
                         total=total,
                         total_expenses=total_expenses,
                         profit=profit)

@app.route('/project/<int:project_id>/save_all', methods=['POST'])
@login_required
def save_all(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('projects'))
    
    try:
        project.name = request.form['name']
        project.address = request.form['address']
        project.status = request.form['status']
        project.client_name = request.form.get('client_name', '')
        project.client_phone = request.form.get('client_phone', '')
        project.actual_end_date = normalize_date(request.form.get('actual_end_date'))
        project.estimate_mode = request.form.get('estimate_mode', project.estimate_mode)
        project.updated_at = datetime.utcnow()
        
        timeline = ProjectTimeline.query.filter_by(project_id=project.id).first()
        start_date = normalize_date(request.form.get('start_date'))
        end_date = normalize_date(request.form.get('end_date'))
        
        if timeline:
            timeline.start_date = start_date or None
            timeline.end_date = end_date or None
        elif start_date or end_date:
            timeline = ProjectTimeline(project_id=project.id, start_date=start_date or None, end_date=end_date or None)
            db.session.add(timeline)
        
        # обновляем существующие работы
        for work in project.works:
            qty_key = f'work_qty_{work.id}'
            if qty_key in request.form:
                work.quantity = float(request.form[qty_key])
                if work.custom_total_price:
                    work.total_price = work.custom_total_price * work.quantity
                else:
                    price = (work.custom_work_price if work.custom_work_price else work.service.work_price)
                    if work.custom_material_price is not None:
                        price += work.custom_material_price
                    else:
                        price += work.service.material_price
                    work.total_price = price * work.quantity
        
        # обновляем существующие расходы
        for expense in project.expenses:
            name_key = f'expense_name_{expense.id}'
            amount_key = f'expense_amount_{expense.id}'
            if name_key in request.form and amount_key in request.form:
                expense.name = request.form[name_key]
                expense.amount = float(request.form[amount_key])
        
        # добавляем новую работу
        if request.form.get('add_work'):
            service_id = request.form.get('new_service_id')
            quantity = request.form.get('new_quantity')
            if service_id and quantity:
                service = Service.query.get(int(service_id))
                qty = float(quantity)
                total_price = (service.work_price + service.material_price) * qty
                db.session.add(ProjectWork(project_id=project.id, service_id=service.id, quantity=qty, total_price=total_price))
        
        # добавляем новый расход
        if request.form.get('add_expense'):
            name = request.form.get('new_expense_name')
            amount = request.form.get('new_expense_amount')
            if name and amount:
                db.session.add(Expense(project_id=project.id, name=name, amount=float(amount)))
        
        # добавляем новую покупку
        if request.form.get('add_purchase'):
            purchase_name = request.form.get('new_purchase_name')
            if purchase_name:
                db.session.add(Purchase(project_id=project.id, name=purchase_name))
        
        # обновляем статус покупок
        for purchase in project.purchases:
            purchase.is_purchased = f'purchased_{purchase.id}' in request.form
        
        db.session.commit()
        flash('Все данные сохранены', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при сохранении: {str(e)}', 'danger')
    
    return redirect(url_for('project_detail', project_id=project_id))

@app.route('/project/<int:project_id>/copy')
@login_required
def copy_project(project_id):
    original = Project.query.get_or_404(project_id)
    if original.user_id != current_user.id:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('projects'))
    
    try:
        new_project = Project(
            name=f"{original.name} (копия)",
            address=original.address,
            status="В работе",
            client_name=original.client_name,
            client_phone=original.client_phone,
            actual_end_date=original.actual_end_date,
            estimate_mode=original.estimate_mode,
            user_id=current_user.id,
            legal_entity_id=original.legal_entity_id
        )
        db.session.add(new_project)
        db.session.commit()
        
        if original.timeline:
            db.session.add(ProjectTimeline(project_id=new_project.id, start_date=original.timeline.start_date, end_date=original.timeline.end_date))
        
        for work in original.works:
            db.session.add(ProjectWork(project_id=new_project.id, service_id=work.service_id, quantity=work.quantity, total_price=work.total_price,
                                     custom_work_price=work.custom_work_price, custom_material_price=work.custom_material_price, 
                                     custom_total_price=work.custom_total_price, custom_name=work.custom_name))
        
        for expense in original.expenses:
            db.session.add(Expense(project_id=new_project.id, name=expense.name, amount=expense.amount))
        
        for purchase in original.purchases:
            db.session.add(Purchase(project_id=new_project.id, name=purchase.name, quantity=purchase.quantity, is_purchased=False))
        
        db.session.commit()
        flash('Проект скопирован', 'success')
        return redirect(url_for('project_detail', project_id=new_project.id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при копировании: {str(e)}', 'danger')
        return redirect(url_for('projects'))

# AJAX маршруты 

@app.route('/project/<int:project_id>/delete_work_ajax/<int:work_id>', methods=['DELETE'])
@login_required
def delete_work_ajax(project_id, work_id):
    work = ProjectWork.query.get_or_404(work_id)
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    db.session.delete(work)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/project/<int:project_id>/delete_expense_ajax/<int:expense_id>', methods=['DELETE'])
@login_required
def delete_expense_ajax(project_id, expense_id):
    expense = Expense.query.get_or_404(expense_id)
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    db.session.delete(expense)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/project/<int:project_id>/delete_purchase_ajax/<int:purchase_id>', methods=['DELETE'])
@login_required
def delete_purchase_ajax(project_id, purchase_id):
    purchase = Purchase.query.get_or_404(purchase_id)
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    db.session.delete(purchase)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if request.method == 'POST':
        Setting.set('company_name', request.form.get('company_name', ''))
        Setting.set('company_phone', request.form.get('company_phone', ''))
        Setting.set('company_email', request.form.get('company_email', ''))
        Setting.set('company_inn', request.form.get('company_inn', ''))
        flash('Настройки сохранены', 'success')
        return redirect(url_for('settings'))
    
    return render_template('settings.html',
                         company_name=Setting.get('company_name', 'Карманный Прораб'),
                         company_phone=Setting.get('company_phone', '+7(XXX)XXX-XX-XX'),
                         company_email=Setting.get('company_email', 'info@karman-prorab.ru'),
                         company_inn=Setting.get('company_inn', ''))

# PDF и экспорт 

@app.route('/project/<int:project_id>/estimate/pdf')
@login_required
def project_estimate_pdf(project_id):
    if not WEASYPRINT_AVAILABLE:
        flash('Генерация PDF временно недоступна. Используйте экспорт в Excel.', 'warning')
        return redirect(url_for('project_detail', project_id=project_id))
    
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('projects'))
    
    try:
        mode = request.args.get('mode', 'full')
        entity_id = request.args.get('entity_id')
        photo_ids_param = request.args.get('photo_ids', '')
        photo_ids = [int(x) for x in photo_ids_param.split(',') if x]
        
        works = ProjectWork.query.filter_by(project_id=project_id).all()
        timeline = ProjectTimeline.query.filter_by(project_id=project_id).first()
        total_income = sum(w.total_price for w in works)
        expenses = Expense.query.filter_by(project_id=project_id).all()
        total_expenses = sum(e.amount for e in expenses)
        profit = total_income - total_expenses
        
        works_list = []
        for i, w in enumerate(works, 1):
            name = w.custom_name if w.custom_name else w.service.name
            unit = w.service.unit
            quantity = w.quantity
            work_price = w.custom_work_price if w.custom_work_price else w.service.work_price
            material_price = w.custom_material_price if w.custom_material_price else w.service.material_price
            works_list.append({
                'num': i, 'name': name, 'unit': unit, 'quantity': quantity,
                'work_price': work_price, 'material_price': material_price, 'total': w.total_price
            })
        
        logo_url = None
        if entity_id:
            entity = LegalEntity.query.get(entity_id)
            if entity and entity.logo_filename and entity.user_id == current_user.id:
                logo_path = os.path.join(app.config['UPLOAD_FOLDER_LOGO'], entity.logo_filename)
                if os.path.exists(logo_path):
                    logo_url = url_for('static', filename=f'uploads/logo/{entity.logo_filename}', _external=True)
        
        photos = ProjectPhoto.query.filter(ProjectPhoto.id.in_(photo_ids)).all() if photo_ids else []
        
        if mode == 'internal':
            template_name = 'pdf_internal.html'
        elif mode == 'short':
            template_name = 'pdf_estimate_short.html'
        else:
            template_name = 'pdf_estimate.html'
        
        html_content = render_template(template_name, 
                                     project=project, works=works_list, total_income=total_income,
                                     timeline=timeline, 
                                     company_name=Setting.get('company_name', 'Карманный Прораб'),
                                     company_phone=Setting.get('company_phone', '+7(XXX)XXX-XX-XX'),
                                     company_email=Setting.get('company_email', 'info@karman-prorab.ru'),
                                     company_inn=Setting.get('company_inn', ''), 
                                     estimate_mode=mode,
                                     logo_url=logo_url,
                                     expenses=expenses,
                                     total_expenses=total_expenses,
                                     profit=profit,
                                     photos=photos)
        
        # Генерация PDF во временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            HTML(string=html_content).write_pdf(tmp.name)
            tmp_path = tmp.name
        
        def cleanup():
            try:
                os.unlink(tmp_path)
            except:
                pass
        
        response = send_file(tmp_path, as_attachment=True, download_name=f'Смета_{project.name}.pdf', mimetype='application/pdf')
        response.call_on_close(cleanup)
        return response
        
    except Exception as e:
        logger.error(f"PDF generation error: {e}")
        flash(f'Ошибка при генерации PDF: {str(e)}', 'danger')
        return redirect(url_for('project_detail', project_id=project_id))

@app.route('/project/<int:project_id>/export_excel')
@login_required
def export_excel(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('projects'))
    
    try:
        works = ProjectWork.query.filter_by(project_id=project_id).all()
        expenses = Expense.query.filter_by(project_id=project_id).all()
        total_income = sum(w.total_price for w in works)
        total_expenses = sum(e.amount for e in expenses)
        profit = total_income - total_expenses
        
        wb = Workbook()
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # лист со сметой
        ws1 = wb.active
        ws1.title = "Смета"
        ws1.merge_cells('A1:F1')
        ws1['A1'] = f'СМЕТА: {project.name}'
        ws1['A1'].font = Font(bold=True, size=14)
        ws1['A1'].alignment = Alignment(horizontal="center")
        
        ws1['A3'] = 'Объект:'
        ws1['B3'] = project.name
        ws1['A4'] = 'Адрес:'
        ws1['B4'] = project.address or '-'
        ws1['A5'] = 'Заказчик:'
        ws1['B5'] = project.client_name or '-'
        
        headers = ['№', 'Наименование работ', 'Ед. изм.', 'Кол-во', 'Цена, руб', 'Сумма, руб']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 9
        for i, work in enumerate(works, 1):
            name = work.custom_name if work.custom_name else work.service.name
            unit = work.service.unit
            price = (work.custom_work_price if work.custom_work_price else work.service.work_price) + (work.custom_material_price if work.custom_material_price else work.service.material_price)
            
            ws1.cell(row=row, column=1, value=i).border = thin_border
            ws1.cell(row=row, column=2, value=name).border = thin_border
            ws1.cell(row=row, column=3, value=unit).border = thin_border
            ws1.cell(row=row, column=4, value=float(work.quantity)).border = thin_border
            ws1.cell(row=row, column=5, value=round(price, 2)).border = thin_border
            ws1.cell(row=row, column=6, value=round(work.total_price, 2)).border = thin_border
            row += 1
        
        total_row = row
        ws1.cell(row=total_row, column=5, value='ИТОГО:').font = Font(bold=True)
        ws1.cell(row=total_row, column=6, value=round(total_income, 2)).font = Font(bold=True)
        for col in range(1, 7):
            ws1.cell(row=total_row, column=col).border = thin_border
        
        # лист с расходами
        ws2 = wb.create_sheet("Расходы")
        ws2['A1'] = f'РАСХОДЫ ПО ПРОЕКТУ: {project.name}'
        ws2['A1'].font = Font(bold=True, size=14)
        
        headers2 = ['№', 'Наименование расхода', 'Сумма, руб']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 4
        for i, expense in enumerate(expenses, 1):
            ws2.cell(row=row, column=1, value=i).border = thin_border
            ws2.cell(row=row, column=2, value=expense.name).border = thin_border
            ws2.cell(row=row, column=3, value=round(expense.amount, 2)).border = thin_border
            row += 1
        
        total_row2 = row
        ws2.cell(row=total_row2, column=2, value='ИТОГО:').font = Font(bold=True)
        ws2.cell(row=total_row2, column=3, value=round(total_expenses, 2)).font = Font(bold=True)
        for col in range(1, 4):
            ws2.cell(row=total_row2, column=col).border = thin_border
        
        # лист с прибылью
        ws3 = wb.create_sheet("Прибыль")
        ws3['A1'] = f'ИТОГОВАЯ ПРИБЫЛЬ: {project.name}'
        ws3['A1'].font = Font(bold=True, size=14)
        
        ws3['A3'] = 'Доход по смете:'
        ws3['B3'] = round(total_income, 2)
        ws3['A4'] = 'Расходы по проекту:'
        ws3['B4'] = round(total_expenses, 2)
        ws3['A6'] = 'ИТОГОВАЯ ПРИБЫЛЬ:'
        ws3['B6'] = round(profit, 2)
        ws3['A6'].font = Font(bold=True, size=12)
        ws3['B6'].font = Font(bold=True, size=12)
        if profit >= 0:
            ws3['B6'].fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        else:
            ws3['B6'].fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        
        # Автоматическое изменение ширины столбцов
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, as_attachment=True, download_name=f'Смета_{project.name}.xlsx', 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        flash(f'Ошибка при экспорте: {str(e)}', 'danger')
        return redirect(url_for('project_detail', project_id=project_id))

# фото чеков 

@app.route('/expense/<int:expense_id>/add_photo', methods=['POST'])
@login_required
def add_expense_photo(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    if expense.project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    if 'photo' not in request.files:
        return jsonify({'success': False, 'message': 'Файл не выбран'})
    
    file = request.files['photo']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Файл не выбран'})
    
    if file and allowed_file(file.filename):
        original_name = secure_filename(file.filename)
        filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{original_name}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER_RECEIPTS'], filename)
        file.save(filepath)
        
        photo = ExpensePhoto(expense_id=expense_id, filename=filename, filepath=filename)
        db.session.add(photo)
        db.session.commit()
        
        return jsonify({'success': True, 'photo_id': photo.id, 'filename': filename})
    
    return jsonify({'success': False, 'message': 'Недопустимый формат файла'})

@app.route('/expense/photo/<int:photo_id>/delete', methods=['DELETE'])
@login_required
def delete_expense_photo(photo_id):
    photo = ExpensePhoto.query.get_or_404(photo_id)
    if photo.expense.project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    file_path = os.path.join(app.config['UPLOAD_FOLDER_RECEIPTS'], photo.filepath)
    if os.path.exists(file_path):
        os.remove(file_path)
    
    db.session.delete(photo)
    db.session.commit()
    return jsonify({'success': True})

# фото проектов 

@app.route('/project/<int:project_id>/add_photo', methods=['POST'])
@login_required
def add_project_photo(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    if 'photo' not in request.files:
        return jsonify({'success': False, 'message': 'Файл не выбран'})
    
    file = request.files['photo']
    photo_type = request.form.get('photo_type', 'progress')
    description = request.form.get('description', '')
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Файл не выбран'})
    
    if file and allowed_file(file.filename):
        original_name = secure_filename(file.filename)
        filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{original_name}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER_PHOTOS'], filename)
        file.save(filepath)
        
        photo = ProjectPhoto(project_id=project_id, filename=filename, filepath=filename, photo_type=photo_type, description=description)
        db.session.add(photo)
        db.session.commit()
        
        return jsonify({'success': True, 'photo_id': photo.id, 'filename': filename})
    
    return jsonify({'success': False, 'message': 'Недопустимый формат файла'})

@app.route('/project/photo/<int:photo_id>/delete', methods=['DELETE'])
@login_required
def delete_project_photo(photo_id):
    photo = ProjectPhoto.query.get_or_404(photo_id)
    if photo.project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    file_path = os.path.join(app.config['UPLOAD_FOLDER_PHOTOS'], photo.filepath)
    if os.path.exists(file_path):
        os.remove(file_path)
    
    db.session.delete(photo)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/project/photo/<int:photo_id>')
@login_required
def get_project_photo(photo_id):
    photo = ProjectPhoto.query.get_or_404(photo_id)
    if photo.project.user_id != current_user.id:
        abort(403)
    
    file_path = os.path.join(app.config['UPLOAD_FOLDER_PHOTOS'], photo.filepath)
    if not os.path.exists(file_path):
        abort(404)
    
    return send_file(file_path)

# AJAX маршруты для проектов

@app.route('/project/<int:project_id>/autosave', methods=['POST'])
@login_required
def autosave(project_id):
    project = Project.query.get_or_404(project_id)
    
    if project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    data = request.get_json()
    field = data.get('field')
    value = data.get('value')
    
    try:
        if field in ['start_date', 'end_date', 'actual_end_date']:
            value = normalize_date(value)
        elif field == 'estimate_mode':
            project.estimate_mode = value
            db.session.commit()
            return jsonify({'success': True})
        
        if field == 'name':
            project.name = value
        elif field == 'address':
            project.address = value
        elif field == 'status':
            project.status = value
        elif field == 'client_name':
            project.client_name = value
        elif field == 'client_phone':
            project.client_phone = value
        elif field == 'actual_end_date':
            project.actual_end_date = value
        elif field == 'start_date':
            timeline = ProjectTimeline.query.filter_by(project_id=project.id).first()
            if not timeline:
                timeline = ProjectTimeline(project_id=project.id)
                db.session.add(timeline)
            timeline.start_date = value
        elif field == 'end_date':
            timeline = ProjectTimeline.query.filter_by(project_id=project.id).first()
            if not timeline:
                timeline = ProjectTimeline(project_id=project.id)
                db.session.add(timeline)
            timeline.end_date = value
        elif field.startswith('work_qty_'):
            work_id = int(field.split('_')[-1])
            work = ProjectWork.query.get(work_id)
            if work and work.project.user_id == current_user.id:
                work.quantity = float(value)
                if work.custom_total_price:
                    work.total_price = work.custom_total_price * work.quantity
                else:
                    price = (work.custom_work_price if work.custom_work_price else work.service.work_price)
                    if work.custom_material_price is not None:
                        price += work.custom_material_price
                    else:
                        price += work.service.material_price
                    work.total_price = price * work.quantity
        elif field.startswith('work_work_price_'):
            work_id = int(field.split('_')[-1])
            work = ProjectWork.query.get(work_id)
            if work and work.project.user_id == current_user.id:
                work.custom_work_price = float(value)
                work.custom_total_price = None
                price = work.custom_work_price
                if work.custom_material_price is not None:
                    price += work.custom_material_price
                else:
                    price += work.service.material_price
                work.total_price = price * work.quantity
        elif field.startswith('work_material_price_'):
            work_id = int(field.split('_')[-1])
            work = ProjectWork.query.get(work_id)
            if work and work.project.user_id == current_user.id:
                work.custom_material_price = float(value)
                work.custom_total_price = None
                price = (work.custom_work_price if work.custom_work_price else work.service.work_price)
                price += work.custom_material_price
                work.total_price = price * work.quantity
        elif field.startswith('work_total_price_'):
            work_id = int(field.split('_')[-1])
            work = ProjectWork.query.get(work_id)
            if work and work.project.user_id == current_user.id:
                work.custom_total_price = float(value)
                work.total_price = work.custom_total_price * work.quantity
        elif field.startswith('work_name_'):
            work_id = int(field.split('_')[-1])
            work = ProjectWork.query.get(work_id)
            if work and work.project.user_id == current_user.id:
                work.custom_name = value
        elif field.startswith('expense_name_'):
            expense_id = int(field.split('_')[-1])
            expense = Expense.query.get(expense_id)
            if expense and expense.project.user_id == current_user.id:
                expense.name = value
        elif field.startswith('expense_amount_'):
            expense_id = int(field.split('_')[-1])
            expense = Expense.query.get(expense_id)
            if expense and expense.project.user_id == current_user.id:
                expense.amount = float(value)
        elif field.startswith('purchased_'):
            purchase_id = int(field.split('_')[-1])
            purchase = Purchase.query.get(purchase_id)
            if purchase and purchase.project.user_id == current_user.id:
                purchase.is_purchased = value == 'true' or value is True
        elif field.startswith('purchase_quantity_'):
            purchase_id = int(field.split('_')[-1])
            purchase = Purchase.query.get(purchase_id)
            if purchase and purchase.project.user_id == current_user.id:
                purchase.quantity = int(value)
        
        project.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Autosave error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/project/<int:project_id>/add_work_ajax', methods=['POST'])
@login_required
def add_work_ajax(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    data = request.get_json()
    service_id = data.get('service_id')
    quantity = float(data.get('quantity', 1))
    
    if not service_id or quantity <= 0:
        return jsonify({'success': False, 'message': 'Неверные данные'}), 400
    
    service = Service.query.get_or_404(service_id)
    total_price = (service.work_price + service.material_price) * quantity
    
    work = ProjectWork(project_id=project_id, service_id=service.id, quantity=quantity, total_price=total_price)
    db.session.add(work)
    db.session.commit()
    
    return jsonify({'success': True, 'work_id': work.id, 'service_name': service.name, 'unit': service.unit,
                    'quantity': quantity, 'work_price': service.work_price, 'material_price': service.material_price, 
                    'total_price': total_price})

@app.route('/project/<int:project_id>/add_expense_ajax', methods=['POST'])
@login_required
def add_expense_ajax(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    data = request.get_json()
    name = data.get('name')
    amount = float(data.get('amount', 0))
    
    if not name or amount < 0:
        return jsonify({'success': False, 'message': 'Неверные данные'}), 400
    
    expense = Expense(project_id=project_id, name=name, amount=amount)
    db.session.add(expense)
    db.session.commit()
    
    return jsonify({'success': True, 'expense_id': expense.id, 'name': expense.name, 'amount': expense.amount})

@app.route('/project/<int:project_id>/add_purchase_ajax', methods=['POST'])
@login_required
def add_purchase_ajax(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    data = request.get_json()
    name = data.get('name')
    quantity = int(data.get('quantity', 1))
    
    if not name:
        return jsonify({'success': False, 'message': 'Неверные данные'}), 400
    
    purchase = Purchase(project_id=project_id, name=name, quantity=quantity)
    db.session.add(purchase)
    db.session.commit()
    
    return jsonify({'success': True, 'purchase_id': purchase.id, 'name': purchase.name, 'quantity': purchase.quantity})

@app.route('/project/<int:project_id>/get_totals')
@login_required
def get_totals(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    works = ProjectWork.query.filter_by(project_id=project_id).all()
    expenses = Expense.query.filter_by(project_id=project_id).all()
    total_income = sum(w.total_price for w in works)
    total_expenses = sum(e.amount for e in expenses)
    profit = total_income - total_expenses
    progress = calculate_progress(project)
    
    return jsonify({'total_income': total_income, 'total_expenses': total_expenses, 'profit': profit, 'progress': progress})

@app.route('/project/<int:project_id>/save_all_ajax', methods=['POST'])
@login_required
def save_all_ajax(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.user_id != current_user.id:
            return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
        
        data = request.json
        
        if 'name' in data:
            project.name = data['name']
        if 'address' in data:
            project.address = data['address']
        if 'status' in data:
            project.status = data['status']
        if 'client_name' in data:
            project.client_name = data['client_name']
        if 'client_phone' in data:
            project.client_phone = data['client_phone']
        if 'actual_end_date' in data and data['actual_end_date']:
            project.actual_end_date = normalize_date(data['actual_end_date'])
        if 'estimate_mode' in data:
            project.estimate_mode = data['estimate_mode']
        
        timeline = ProjectTimeline.query.filter_by(project_id=project.id).first()
        if not timeline:
            timeline = ProjectTimeline(project_id=project.id)
            db.session.add(timeline)
        
        if 'start_date' in data and data['start_date']:
            timeline.start_date = normalize_date(data['start_date'])
        if 'end_date' in data and data['end_date']:
            timeline.end_date = normalize_date(data['end_date'])
        
        project.updated_at = datetime.utcnow()
        db.session.commit()
        
        progress = calculate_progress(project)
        return jsonify({"success": True, "progress": progress, "message": "Сохранено успешно"})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Save all ajax error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/search-projects')
@login_required
def search_projects_api():
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify([])
    
    projects = Project.query.filter(
        Project.user_id == current_user.id,
        (Project.name.ilike(f'%{query}%') | 
         Project.address.ilike(f'%{query}%') |
         Project.client_name.ilike(f'%{query}%'))
    ).limit(10).all()
    
    return jsonify([{
        'id': p.id,
        'name': p.name,
        'address': p.address or '',
        'client_name': p.client_name or ''
    } for p in projects])

# shablons (temlates) upload

@app.route('/project/<int:project_id>/upload_template', methods=['POST'])
@login_required
def upload_template(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Доступ запрещен'}), 403
    
    template_name = request.form.get('template_name')
    file = request.files.get('template_file')
    apply_to_project = request.form.get('apply_to_project') == 'on'
    
    if not file or not template_name:
        return jsonify({'success': False, 'message': 'Название шаблона и файл обязательны'})
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': 'Недопустимый формат файла. Поддерживаются .xlsx и .csv'})
    
    try:
        original_name = secure_filename(file.filename)
        filename = f"template_{current_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{original_name}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER_TEMPLATES'], filename)
        file.save(filepath)
        
        template = Template(
            user_id=current_user.id,
            name=template_name,
            filename=filename
        )
        db.session.add(template)
        db.session.commit()
        
        added_count = 0
        if apply_to_project:
            wb = load_workbook(filepath)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1] or not row[2]:
                    continue
                
                service_name = str(row[0]).strip()
                unit = str(row[1]).strip()
                price = float(row[2])
                
                existing_service = Service.query.filter_by(
                    name=service_name, 
                    unit=unit,
                    work_price=price
                ).first()
                
                if not existing_service:
                    existing_service = Service(
                        name=service_name,
                        unit=unit,
                        work_price=price,
                        material_price=0
                    )
                    db.session.add(existing_service)
                    db.session.commit()
                
                work = ProjectWork(
                    project_id=project.id,
                    service_id=existing_service.id,
                    quantity=1,
                    total_price=price
                )
                db.session.add(work)
                added_count += 1
            
            db.session.commit()
        
        return jsonify({'success': True, 'applied': apply_to_project, 'added_count': added_count})
        
    except Exception as e:
        logger.error(f"Template upload error: {e}")
        return jsonify({'success': False, 'message': f'Ошибка: {str(e)}'})

# user loader

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# инициализация бд

def init_db():
    """Инициализация базы данных с проверкой существования"""
    with app.app_context():
        db.create_all()
        logger.info(f"Database initialized at {DB_PATH}")
        
        # Проверяем, есть ли администратор (опционально)
        if not User.query.filter_by(email='admin@example.com').first():
            logger.info("No admin user found, you can register first user")

# запуск

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    
    logger.info(f"Starting Karman Prorab on port {port}, debug={debug}")
    logger.info(f"Template folder: {TEMPLATE_DIR}")
    logger.info(f"Static folder: {STATIC_DIR}")
    logger.info(f"Database path: {DB_PATH}")
    
    app.run(host='0.0.0.0', port=port, debug=debug)